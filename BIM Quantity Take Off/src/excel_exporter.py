"""
Excel Export Module
Exports BOQ data to professionally formatted Excel files.
"""

import logging
from pathlib import Path
from typing import Optional, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports BOQ data to Excel with professional formatting.
    
    Creates well-formatted Excel workbooks suitable for professional
    civil engineering documentation and cost estimation.
    """
    
    def __init__(self, output_path: str):
        """
        Initialize the Excel exporter.
        
        Args:
            output_path: Path where the Excel file will be saved
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
    
    def export_boq(
        self,
        boq_df: pd.DataFrame,
        summary: Optional[Dict[str, Any]] = None,
        project_info: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Export BOQ DataFrame to Excel with formatting.
        Creates one sheet per element category (type) as per requirements.
        
        Args:
            boq_df: BOQ DataFrame
            summary: Optional summary dictionary
            project_info: Optional project information dictionary
        """
        # Ensure we always have at least one sheet
        sheets_created = False
        
        # Create Excel writer
        with pd.ExcelWriter(str(self.output_path), engine='openpyxl') as writer:
            # Access workbook for formatting (available after first write)
            workbook = None
            
            # Create one sheet per element category (type)
            if not boq_df.empty and 'element_type' in boq_df.columns:
                element_types = boq_df['element_type'].unique()
                
                if len(element_types) > 0:
                    for elem_type in element_types:
                        # Filter BOQ for this element type
                        type_df = boq_df[boq_df['element_type'] == elem_type].copy()
                        
                        # Clean sheet name (Excel sheet names have limitations)
                        sheet_name = self._clean_sheet_name(elem_type)
                        
                        # Add item numbers if not already added
                        if 'item_no' in type_df.columns:
                             type_df = type_df.drop(columns=['item_no'])
                        type_df.insert(0, 'item_no', range(1, len(type_df) + 1))
                        
                        # Write to sheet
                        type_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheets_created = True
            
            # Fallback: single sheet if no element_type column or empty DataFrame
            if not sheets_created:
                if boq_df.empty:
                    # Create empty DataFrame with proper columns
                    boq_df = pd.DataFrame(columns=['item_no', 'element_type', 'description', 
                                                   'unit', 'quantity', 'storey', 'material'])
                    logger.warning("BOQ DataFrame is empty. Creating empty sheet.")
                
                # Add item numbers if not already added
                if not boq_df.empty:
                    if 'item_no' in boq_df.columns:
                         boq_df = boq_df.drop(columns=['item_no'])
                    boq_df.insert(0, 'item_no', range(1, len(boq_df) + 1))
                
                boq_df.to_excel(writer, sheet_name='Bill of Quantities', index=False)
                sheets_created = True
            
            # Write summary sheet if provided
            if summary:
                summary_df = self._create_summary_dataframe(summary)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                sheets_created = True
            
            # Write project info sheet if provided
            if project_info:
                project_df = self._create_project_info_dataframe(project_info)
                project_df.to_excel(writer, sheet_name='Project Information', index=False)
                sheets_created = True
            
            # Ensure at least one sheet exists (fallback)
            if not sheets_created:
                # Create a minimal placeholder sheet
                placeholder_df = pd.DataFrame({
                    'Message': ['No data found in IFC file. Please check the file contains civil engineering elements (Beam, Column, Slab, Wall, Footing).']
                })
                placeholder_df.to_excel(writer, sheet_name='No Data', index=False)
            
            # Format the Excel file before closing (using workbook object from writer)
            try:
                workbook = writer.book  # Get workbook after sheets are written
                if workbook:
                    self._format_excel_file_workbook(workbook, boq_df, summary, project_info)
            except Exception as e:
                logger.warning(f"Error formatting Excel file: {e}. File will be saved without formatting.")
        
        logger.info(f"BOQ exported successfully to: {self.output_path}")
        
        logger.info(f"BOQ exported successfully to: {self.output_path}")
    
    def _clean_sheet_name(self, name: str) -> str:
        """
        Clean sheet name to meet Excel requirements.
        Excel sheet names must be <= 31 characters and cannot contain certain characters.
        
        Args:
            name: Original sheet name
            
        Returns:
            Cleaned sheet name
        """
        # Remove 'Ifc' prefix
        clean_name = name.replace('Ifc', '') if name.startswith('Ifc') else name
        
        # Remove invalid characters: / \ ? * [ ]
        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        # Truncate to 31 characters (Excel limit)
        if len(clean_name) > 31:
            clean_name = clean_name[:31]
        
        # Ensure it's not empty
        if not clean_name:
            clean_name = "Sheet1"
        
        return clean_name
    
    def _create_summary_dataframe(self, summary: Dict[str, Any]) -> pd.DataFrame:
        """Create a DataFrame from summary dictionary."""
        data = []
        for key, value in summary.items():
            if isinstance(value, list):
                value = ', '.join(str(v) for v in value)
            data.append({'Metric': key.replace('_', ' ').title(), 'Value': value})
        return pd.DataFrame(data)
    
    def _create_project_info_dataframe(self, project_info: Dict[str, Any]) -> pd.DataFrame:
        """Create a DataFrame from project info dictionary."""
        data = []
        for key, value in project_info.items():
            data.append({'Property': key.replace('_', ' ').title(), 'Value': str(value)})
        return pd.DataFrame(data)
    
    def _format_excel_file_workbook(
        self,
        workbook: Any,
        boq_df: pd.DataFrame,
        summary: Optional[Dict[str, Any]],
        project_info: Optional[Dict[str, Any]]
    ) -> None:
        """Apply professional formatting to the Excel workbook (using workbook object directly)."""
        # Format all BOQ sheets (one per element type)
        for sheet_name in workbook.sheetnames:
            if sheet_name not in ['Summary', 'Project Information', 'No Data']:
                # This is a BOQ sheet
                sheet = workbook[sheet_name]
                self._format_boq_sheet(sheet, boq_df)
            elif sheet_name == 'No Data':
                # Format the placeholder sheet
                sheet = workbook[sheet_name]
                self._format_info_sheet(sheet)
        
        # Format summary sheet
        if summary and 'Summary' in workbook.sheetnames:
            self._format_summary_sheet(workbook['Summary'])
        
        # Format project info sheet
        if project_info and 'Project Information' in workbook.sheetnames:
            self._format_info_sheet(workbook['Project Information'])
    
    def _format_excel_file(
        self,
        boq_df: pd.DataFrame,
        summary: Optional[Dict[str, Any]],
        project_info: Optional[Dict[str, Any]]
    ) -> None:
        """Apply professional formatting to the Excel file (fallback method)."""
        try:
            workbook = load_workbook(str(self.output_path))
            
            # Format all BOQ sheets (one per element type)
            for sheet_name in workbook.sheetnames:
                if sheet_name not in ['Summary', 'Project Information', 'No Data']:
                    # This is a BOQ sheet
                    sheet = workbook[sheet_name]
                    self._format_boq_sheet(sheet, boq_df)
                elif sheet_name == 'No Data':
                    # Format the placeholder sheet
                    sheet = workbook[sheet_name]
                    self._format_info_sheet(sheet)
            
            # Format summary sheet
            if summary and 'Summary' in workbook.sheetnames:
                self._format_summary_sheet(workbook['Summary'])
            
            # Format project info sheet
            if project_info and 'Project Information' in workbook.sheetnames:
                self._format_info_sheet(workbook['Project Information'])
            
            workbook.save(str(self.output_path))
        except Exception as e:
            logger.warning(f"Could not format Excel file after saving: {e}")
    
    def _format_boq_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """Format the BOQ worksheet with clean formatting."""
        if worksheet.max_row == 0:
            return
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Identify numeric columns (Quantity and other numeric fields)
        numeric_columns = []
        if not df.empty:
            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                # Check if column contains numeric data
                if col.lower() in ['quantity', 'volume_m3', 'area_m2', 'length_m', 'count']:
                    numeric_columns.append(col_letter)
                elif df[col].dtype in ['int64', 'float64']:
                    numeric_columns.append(col_letter)
                    
        # Identify currency columns
        currency_columns = []
        if not df.empty:
            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                if col in ['Rate', 'Total Cost']:
                    currency_columns.append(col_letter)
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                # Right-align numeric columns
                if cell.column_letter in numeric_columns:
                    try:
                        if cell.value is not None:
                            float(cell.value)
                            cell.alignment = right_align
                            # Format numbers
                            if cell.column_letter in currency_columns:
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '0.000'
                    except (ValueError, TypeError):
                        cell.alignment = left_align
                else:
                    cell.alignment = left_align
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        if worksheet.max_row > 1:
            worksheet.freeze_panes = 'A2'
    
    def _format_summary_sheet(self, worksheet) -> None:
        """Format the summary worksheet."""
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_info_sheet(self, worksheet) -> None:
        """Format the project information worksheet."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
